# -*- coding: utf-8 -*-
# 可参考《利用Python下载文件》
# https://blog.csdn.net/sinat_36246371/article/details/62426444
__author__ = 'vita'

from urllib import request
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup

import fun
import json
import requests


# 获取数据
# get方式获取数据
def get_data(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'
    }
    req = request.Request(url, headers=headers)
    response = request.urlopen(req)
    if response.getcode() == 200:
        result = response.read()
        # print(type(result))  # bytes类型
        return result


# 处理数据
def parse_data(html):
    #  创建BeautifulSoup实例，解析html数据
    bs = BeautifulSoup(html, 'html.parser')  # 指定使用html解析器parser

    soup = bs.select('#news_conent_two_text')
    pubDate = bs.select('.news_conent_two_js')[0].find_all('span')[3].text.split('：')[1]
    result = []
    for link in soup[0].find_all('a'):
        print(link.text, link.get('href'))
        row = {
            'title': link.text,
            'url_detail': link.get('href'),
            'pubDate': pubDate
        }
        result.append(row)
    return result


# 存储数据到Excel
def save_to_excel(data, file_name):
    # 创建工作簿Workbook
    book = Workbook()

    # 创建工作表Sheet
    sheet = book.create_sheet(file_name, 0)

    # 向工作表中添加数据
    sheet.append(['title', 'url', 'pubDate'])
    for item in data:
        row = [item['title'], item['url_detail'], item['pubDate']]
        sheet.append(row)

    # 输出保存
    book.save(file_name + '.xlsx')


if __name__ == '__main__':
    host_name = 'http://www.chinaskills-jsw.org'
    # url = 'http://www.chinaskills-jsw.org/content.jsp?id=2c9080b46254d2d10162a66161730177&classid=ff8080814ead5a970151265649470341'
    # result = []
    # # 存储到EXCEL中
    # data = get_data(url)
    # # parse_data(data)
    file_name = '全国职业院校技能大赛赛项赛卷（赛题库）'
    # save_to_excel(parse_data(data), file_name)
    #
    # # 建立存储文件夹
    path = file_name
    fun.mkdir(path)

    # 从Excel中读出url
    book = load_workbook(file_name + '.xlsx')
    sheet = book.worksheets[0]
    rows = list(sheet.rows)
    rows.remove(rows[0])
    for row in rows:
        url_detail = host_name + row[1].value
        store_path = path + '/' + row[0].value
        # print(url_detail, store_path)
        # 下载文件并存储到文件夹中
        try:
           fun.file_download(store_path,url_detail)
        except Exception as e:
            print('Exception:', e)
